import openpyxl
import random
import string

# Generate password with specified length 
def generate_random_password(length=8):
    characters = string.ascii_letters + string.digits
    password = ''.join(random.choice(characters) for _ in range(length))
    return password

# Read Excel file, locate username and password column
# Update passwords for rows that have no existing password
def update_passwords(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active

    username_column = None
    password_column = None

    for column_index, column in enumerate(sheet.iter_cols(min_row=1, max_row=1)):
        if column[0].value == "Username":
            username_column = column_index
        elif column[0].value == "Password":
            password_column = column_index

    # if the column for username or password not found
    if username_column is None or password_column is None:
        print("Username or Password column not found.")
        return

    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        username = row[username_column]
        current_password = row[password_column]

        # If no password is found
        if current_password is None:
            new_password = generate_random_password()
            sheet.cell(row=row_index, column=password_column + 1, value=new_password)

    wb.save(filename)
    print("Passwords updated successfully.")

if __name__ == "__main__":
    # Replace the file path with your exact Excel file path
    excel_filename = "C:\\Users\\user\\OneDrive\\Desktop\\SAINS\\Jenkins-Installer-Automation\\UserList.xlsx"
    update_passwords(excel_filename)

